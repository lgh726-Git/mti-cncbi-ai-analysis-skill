#!/usr/bin/env python3
"""MTI / CNCBI UST weekly P&L analysis (skill entrypoint).

Config resolution order:
  1) CLI --config
  2) $MTI_AI_CONFIG
  3) ~/.config/mti-ai-analysis/config.json
  4) Built-in defaults (local CODEX + Drobo paths)

Writes dated package:
  <ai_root>/AI分析-YYYYMMDD/{.md,.html,.pdf,.json,README.txt}
"""

from __future__ import annotations

import argparse
import json
import os
import shutil
import subprocess
import traceback
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

SKILL_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKDIR = Path.home() / "AI" / "CODEX" / "cncbi-holdings"
DEFAULT_LOG_DIR = Path.home() / "Library" / "Logs" / "mti-ai-analysis"
DEFAULT_CONFIG = Path.home() / ".config" / "mti-ai-analysis" / "config.json"

# Structural economics (stable conversion inputs)
UC08_FACE = 4_299_600.0
UE63_FACE = 1_040_000.0
UC08_MATURITY = date(2054, 8, 15)
UE63_MATURITY = date(2054, 11, 15)
UC08_COUPON = 0.0425
UE63_COUPON = 0.0450

UC08_NOTES = [
    {"code": "WBG SN RCN 5", "notional": 3_060_000.0, "residual": 18_880.20},
    {"code": "WBG RCN 3", "notional": 1_000_000.0, "residual": 519.50},
]
UE63_NOTE = {"code": "WBG SN RCN 6", "notional": 1_000_000.0, "residual": 6_194.00}

BASE_UC08_COUPONS = [
    ("2025-02-04", 41310.00, "RCN5结构票息"),
    ("2025-05-01", 41310.00, "RCN5结构票息"),
    ("2025-08-18", 48248.55, "UC08国债利息"),
    ("2025-10-17", 20677.95, "UC08国债利息"),
    ("2024-12-03", 13992.78, "RCN3结构票息"),
    ("2025-03-04", 14150.00, "RCN3结构票息"),
    ("2025-06-03", 14150.00, "RCN3结构票息"),
    ("2025-09-04", 14150.00, "RCN3结构票息"),
    ("2026-02-20", 91366.50, "UC08国债利息"),
]
BASE_UE63_COUPONS = [
    ("2025-03-04", 14150.00, "RCN6结构票息"),
    ("2025-06-06", 28350.00, "RCN6结构票息"),
    ("2025-11-18", 23400.00, "UE63国债利息"),
    ("2026-05-15", 23400.00, "UE63国债利息"),
]
CLOSED_UG12_1MM = {
    "notional": 1_000_000.0,
    "residual": 874.80,
    "sell_mv": 1_024_695.86,
    "coupons": 31_000.00,
    "note": "2025-09-10 卖出@98",
}
CLOSED_UG12_05 = {
    "notional": 500_000.0,
    "principal_back": 500_000.0,
    "coupons": 8_125.00,
    "note": "SN36 到期还本假设",
}
FALLBACK_Y30 = 0.05181


def money(x: float | None) -> str:
    if x is None:
        return ""
    s = f"{abs(x):,.2f}"
    return f"({s})" if x < 0 else s


def bond_clean(coupon_annual: float, ytm: float, years: float, freq: int = 2, face: float = 100.0) -> float:
    c = coupon_annual / freq * face
    r = ytm / freq
    n = max(1, int(round(years * freq)))
    if abs(r) < 1e-12:
        return face
    return c * (1 - (1 + r) ** (-n)) / r + face / (1 + r) ** n


def years_to(mat: date, asof: date) -> float:
    return max(0.01, (mat - asof).days / 365.25)


def load_config(path: Path | None) -> dict[str, Any]:
    cfg: dict[str, Any] = {
        "workdir": str(DEFAULT_WORKDIR),
        "log_dir": str(DEFAULT_LOG_DIR),
        "ai_root": "",
        "mti_finance_root": "",
        "detail_workbook_candidates": [
            str(Path.home() / "Desktop" / "待解密" / "被兑换债权的理财明细.xlsx"),
        ],
        "mount_script": str(Path.home() / "Library" / "Scripts" / "mount-drobo-smb.sh"),
        "account": "CNCBI One Account 735-3-60028-4-88",
        "entity": "MERIT TALENT INT'L LIMITED",
        "fallback_y30": FALLBACK_Y30,
    }
    candidates = []
    if path:
        candidates.append(Path(path).expanduser())
    env = os.environ.get("MTI_AI_CONFIG")
    if env:
        candidates.append(Path(env).expanduser())
    candidates.append(DEFAULT_CONFIG)
    for p in candidates:
        if p and p.exists():
            try:
                user = json.loads(p.read_text(encoding="utf-8"))
                if isinstance(user, dict):
                    cfg.update(user)
                    cfg["_config_path"] = str(p)
                    break
            except Exception:  # noqa: BLE001
                pass
    return cfg


class Logger:
    def __init__(self, log_dir: Path):
        self.log_dir = log_dir
        self.log_dir.mkdir(parents=True, exist_ok=True)

    def log(self, msg: str) -> None:
        line = f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {msg}"
        print(line)
        with (self.log_dir / "weekly.log").open("a", encoding="utf-8") as f:
            f.write(line + "\n")


def fetch_y30(fallback: float) -> tuple[float, str]:
    import urllib.request

    urls = [
        ("https://query1.finance.yahoo.com/v8/finance/chart/%5ETYX?range=5d&interval=1d", "^TYX"),
        ("https://query1.finance.yahoo.com/v8/finance/chart/%5ETNX?range=5d&interval=1d", "^TNX"),
    ]
    headers = {"User-Agent": "Mozilla/5.0 MTI-AI-Analysis-Skill/1.0"}
    last_err: Exception | None = None
    for url, sym in urls:
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=20) as resp:
                data = json.loads(resp.read().decode())
            px = float(data["chart"]["result"][0]["meta"]["regularMarketPrice"])
            y = px / 100.0
            if sym == "^TNX":
                y = y + 0.0015
                return y, f"Yahoo {sym}={px:.3f}% +15bp→30Y proxy {y*100:.3f}%"
            return y, f"Yahoo {sym}={px:.3f}% → {y*100:.3f}%"
        except Exception as e:  # noqa: BLE001
            last_err = e
    return float(fallback), f"fallback constant {float(fallback)*100:.3f}% ({last_err})"


def ensure_ai_root(cfg: dict[str, Any], logger: Logger) -> Path:
    ai_root = cfg.get("ai_root") or ""
    finance = cfg.get("mti_finance_root") or ""
    if ai_root:
        p = Path(ai_root).expanduser()
    elif finance:
        p = Path(finance).expanduser() / "银行理财" / "AI分析"
    else:
        # local fallback
        p = Path(cfg["workdir"]).expanduser() / "AI分析"

    if not p.parent.exists():
        mount_sh = Path(cfg.get("mount_script") or "").expanduser()
        if mount_sh.exists():
            logger.log(f"path missing; attempting mount: {mount_sh}")
            try:
                subprocess.run(["/bin/zsh", str(mount_sh)], check=False, timeout=120)
            except Exception as e:  # noqa: BLE001
                logger.log(f"mount failed: {e}")
    p.mkdir(parents=True, exist_ok=True)
    return p


def load_detail_workbook(cfg: dict[str, Any]) -> dict[str, Any]:
    if load_workbook is None:
        return {"ok": False, "error": "openpyxl missing (pip install openpyxl)"}
    cands = [Path(x).expanduser() for x in cfg.get("detail_workbook_candidates") or []]
    finance = cfg.get("mti_finance_root")
    if finance:
        cands.append(Path(finance).expanduser() / "银行理财" / "被兑换的国债" / "被兑换债权的理财明细.xlsx")
    for p in cands:
        if not p.exists():
            continue
        head = p.read_bytes()[:8]
        if head[:2] != b"PK":
            return {"ok": False, "error": f"still encrypted: {p}", "path": str(p)}
        try:
            wb = load_workbook(p, data_only=True)
            sheets = {}
            for sn in wb.sheetnames:
                rows = []
                ws = wb[sn]
                for row in ws.iter_rows(values_only=True):
                    if any(v is not None and str(v).strip() != "" for v in row):
                        rows.append([("" if v is None else v) for v in row])
                sheets[sn] = rows
            return {"ok": True, "path": str(p), "sheets": list(sheets.keys())}
        except Exception as e:  # noqa: BLE001
            return {"ok": False, "error": str(e), "path": str(p)}
    return {"ok": False, "error": "detail workbook not found"}


def compute(asof: date, y30: float) -> dict[str, Any]:
    uc08_net_cost = sum(n["notional"] - n["residual"] for n in UC08_NOTES)
    ue63_net_cost = UE63_NOTE["notional"] - UE63_NOTE["residual"]

    px_uc08 = bond_clean(UC08_COUPON, y30, years_to(UC08_MATURITY, asof))
    px_ue63 = bond_clean(UE63_COUPON, y30, years_to(UE63_MATURITY, asof))
    mv_uc08 = UC08_FACE * px_uc08 / 100
    mv_ue63 = UE63_FACE * px_ue63 / 100

    uc08_cpn = sum(a for _, a, _ in BASE_UC08_COUPONS)
    ue63_cpn = sum(a for _, a, _ in BASE_UE63_COUPONS)

    next_uc08 = round(UC08_FACE * UC08_COUPON / 2, 2)
    pending = []
    for y in range(2026, asof.year + 1):
        for md in ((2, 15), (8, 15)):
            d = date(y, md[0], md[1])
            if d <= asof and d >= date(2026, 8, 15):
                if not any(
                    abs((date.fromisoformat(c[0]) - d).days) <= 10 for c in BASE_UC08_COUPONS
                ):
                    pending.append(
                        {
                            "bond": "UC08",
                            "date": d.isoformat(),
                            "amount": next_uc08,
                            "note": "理论票息，待结单确认，未计入综合盈亏",
                        }
                    )

    uc08_price_pnl = mv_uc08 - uc08_net_cost
    ue63_price_pnl = mv_ue63 - ue63_net_cost
    uc08_total = uc08_price_pnl + uc08_cpn
    ue63_total = ue63_price_pnl + ue63_cpn

    ug12a_net = CLOSED_UG12_1MM["notional"] - CLOSED_UG12_1MM["residual"]
    ug12a_total = CLOSED_UG12_1MM["sell_mv"] + CLOSED_UG12_1MM["coupons"] - ug12a_net
    ug12b_total = (
        CLOSED_UG12_05["principal_back"] + CLOSED_UG12_05["coupons"] - CLOSED_UG12_05["notional"]
    )

    open_price = uc08_price_pnl + ue63_price_pnl
    open_cpn = uc08_cpn + ue63_cpn
    open_total = uc08_total + ue63_total
    closed_total = ug12a_total + ug12b_total

    return {
        "asof": asof.isoformat(),
        "y30": y30,
        "open": {
            "uc08": {
                "face": UC08_FACE,
                "net_cost": round(uc08_net_cost, 2),
                "px": round(px_uc08, 3),
                "mv": round(mv_uc08, 2),
                "coupons": round(uc08_cpn, 2),
                "price_pnl": round(uc08_price_pnl, 2),
                "total_pnl": round(uc08_total, 2),
                "coupon_detail": [
                    {"date": d, "amount": a, "type": t} for d, a, t in BASE_UC08_COUPONS
                ],
            },
            "ue63": {
                "face": UE63_FACE,
                "net_cost": round(ue63_net_cost, 2),
                "px": round(px_ue63, 3),
                "mv": round(mv_ue63, 2),
                "coupons": round(ue63_cpn, 2),
                "price_pnl": round(ue63_price_pnl, 2),
                "total_pnl": round(ue63_total, 2),
                "coupon_detail": [
                    {"date": d, "amount": a, "type": t} for d, a, t in BASE_UE63_COUPONS
                ],
            },
        },
        "closed": {
            "ug12_1mm_total": round(ug12a_total, 2),
            "ug12_05_total": round(ug12b_total, 2),
            "total": round(closed_total, 2),
        },
        "pending_coupons": pending,
        "summary": {
            "open_mv": round(mv_uc08 + mv_ue63, 2),
            "open_net_cost": round(uc08_net_cost + ue63_net_cost, 2),
            "open_price_pnl": round(open_price, 2),
            "open_coupons": round(open_cpn, 2),
            "open_total_pnl": round(open_total, 2),
            "closed_total_pnl": round(closed_total, 2),
            "all_in_pnl": round(open_total + closed_total, 2),
        },
    }


def render_html(report: dict[str, Any], y_src: str, cfg: dict[str, Any], detail: dict[str, Any]) -> str:
    s = report["summary"]
    o = report["open"]
    cls = lambda x: "neg" if x < 0 else "pos"  # noqa: E731
    pending_rows = ""
    for p in report.get("pending_coupons", []):
        pending_rows += (
            f"<tr style='background:#fff3cd'><td class='l'>{p['bond']}</td><td class='c'>{p['date']}</td>"
            f"<td>{money(p['amount'])}</td><td class='l note'>{p['note']}</td></tr>"
        )
    if not pending_rows:
        pending_rows = "<tr><td class='l' colspan='4'>无</td></tr>"

    cpn_rows = ""
    for bkey, label in (("uc08", "UC08"), ("ue63", "UE63")):
        for c in sorted(o[bkey]["coupon_detail"], key=lambda x: x["date"]):
            cpn_rows += (
                f"<tr><td class='l'>{label}</td><td class='c'>{c['date']}</td>"
                f"<td class='l'>{c['type']}</td><td class='pos'>{money(c['amount'])}</td></tr>"
            )

    detail_note = (
        f"明细表：{detail.get('path', 'n/a')} "
        f"({'OK' if detail.get('ok') else detail.get('error')})"
    )
    entity = cfg.get("entity", "MERIT TALENT INT'L LIMITED")
    account = cfg.get("account", "CNCBI One Account")

    return f"""<!DOCTYPE html>
<html lang="zh-CN"><head><meta charset="utf-8"/>
<title>AI分析-{report['asof']} MTI美债盈亏（含票息）</title>
<style>
body{{font-family:-apple-system,BlinkMacSystemFont,"PingFang SC",sans-serif;margin:28px;color:#1a1a1a}}
h1{{font-size:22px;margin:0 0 6px}} h2{{font-size:16px;margin:22px 0 8px;border-left:4px solid #1f4e79;padding-left:8px}}
.meta{{color:#555;font-size:13px;margin-bottom:14px}}
table{{border-collapse:collapse;width:100%;font-size:12.5px;margin-bottom:12px}}
th,td{{border:1px solid #d0d7de;padding:6px 8px;text-align:right}}
th{{background:#1f4e79;color:#fff;text-align:center}}
td.l,th.l{{text-align:left}} td.c{{text-align:center}}
.cat-bond{{background:#e8f1fb}} .pos{{color:#0a7a32;font-weight:600}} .neg{{color:#c0392b;font-weight:600}}
.box{{background:#f6f8fa;border:1px solid #d0d7de;border-radius:8px;padding:12px 14px;margin:10px 0 16px}}
.kpi{{display:inline-block;min-width:170px;margin:6px 16px 6px 0}} .kpi b{{display:block;font-size:18px}}
.note{{font-size:12px;color:#555}} .foot{{font-size:11px;color:#666;margin-top:18px;line-height:1.5}}
</style></head><body>
<h1>AI分析-{report['asof'].replace('-', '')} — MTI 美债盈亏（含票息）</h1>
<div class="meta">主体：{entity} &nbsp;|&nbsp; 账户：{account}<br/>
估值日：{report['asof']} &nbsp;|&nbsp; US30Y：{report['y30']*100:.3f}%（{y_src}）<br/>
{detail_note}</div>
<div class="box">
  <div class="kpi">在持市值<br/><b>${money(s['open_mv'])}</b></div>
  <div class="kpi">在持净成本<br/><b>${money(s['open_net_cost'])}</b></div>
  <div class="kpi">价格盈亏<br/><b class="{cls(s['open_price_pnl'])}">${money(s['open_price_pnl'])}</b></div>
  <div class="kpi">已收票息<br/><b class="pos">${money(s['open_coupons'])}</b></div>
  <div class="kpi">在持综合<br/><b class="{cls(s['open_total_pnl'])}">${money(s['open_total_pnl'])}</b></div>
  <div class="kpi">已结清综合<br/><b class="{cls(s['closed_total_pnl'])}">${money(s['closed_total_pnl'])}</b></div>
  <div class="kpi">全链路综合<br/><b class="{cls(s['all_in_pnl'])}">${money(s['all_in_pnl'])}</b></div>
</div>
<h2>在持美债现券</h2>
<table>
<tr><th class="l">品种</th><th>面值</th><th>净成本</th><th>估值价</th><th>市值</th><th>价格盈亏</th><th>已收票息</th><th>综合盈亏</th></tr>
<tr class="cat-bond"><td class="l">UST 4.25% 15Aug2054 (UC08)</td>
<td>{money(o['uc08']['face'])}</td><td>{money(o['uc08']['net_cost'])}</td><td>{o['uc08']['px']:.3f}</td>
<td>{money(o['uc08']['mv'])}</td><td class="{cls(o['uc08']['price_pnl'])}">{money(o['uc08']['price_pnl'])}</td>
<td class="pos">{money(o['uc08']['coupons'])}</td><td class="{cls(o['uc08']['total_pnl'])}">{money(o['uc08']['total_pnl'])}</td></tr>
<tr class="cat-bond"><td class="l">UST 4.50% 15Nov2054 (UE63)</td>
<td>{money(o['ue63']['face'])}</td><td>{money(o['ue63']['net_cost'])}</td><td>{o['ue63']['px']:.3f}</td>
<td>{money(o['ue63']['mv'])}</td><td class="{cls(o['ue63']['price_pnl'])}">{money(o['ue63']['price_pnl'])}</td>
<td class="pos">{money(o['ue63']['coupons'])}</td><td class="{cls(o['ue63']['total_pnl'])}">{money(o['ue63']['total_pnl'])}</td></tr>
<tr style="font-weight:700;background:#dfeaf7"><td class="l">在持小计</td><td>{money(o['uc08']['face']+o['ue63']['face'])}</td>
<td>{money(s['open_net_cost'])}</td><td></td><td>{money(s['open_mv'])}</td>
<td class="{cls(s['open_price_pnl'])}">{money(s['open_price_pnl'])}</td>
<td class="pos">{money(s['open_coupons'])}</td>
<td class="{cls(s['open_total_pnl'])}">{money(s['open_total_pnl'])}</td></tr>
</table>
<div class="note">净成本 = 结构票据本金 − 实物交割退回零头。价格为 Clean price 模型（YTM=US30Y）。</div>
<h2>已结清链路</h2>
<table>
<tr><th class="l">项目</th><th>综合盈亏</th><th class="l">说明</th></tr>
<tr><td class="l">UG12 / RCN8 100万（已卖出）</td><td class="pos">{money(report['closed']['ug12_1mm_total'])}</td><td class="l note">{CLOSED_UG12_1MM['note']}</td></tr>
<tr><td class="l">SN36 50万</td><td class="pos">{money(report['closed']['ug12_05_total'])}</td><td class="l note">{CLOSED_UG12_05['note']}</td></tr>
<tr style="font-weight:700;background:#ead9ff"><td class="l">已结清小计</td><td class="{cls(s['closed_total_pnl'])}">{money(s['closed_total_pnl'])}</td><td></td></tr>
</table>
<h2>票息明细（已计入）</h2>
<table>
<tr><th class="l">债券</th><th class="c">日期</th><th class="l">类型</th><th>金额 USD</th></tr>
{cpn_rows}
<tr style="font-weight:700;background:#e8f5e9"><td class="l" colspan="3">在持票息合计</td><td class="pos">{money(s['open_coupons'])}</td></tr>
</table>
<h2>待确认票息</h2>
<table>
<tr><th class="l">债券</th><th class="c">日期</th><th>金额</th><th class="l">备注</th></tr>
{pending_rows}
</table>
<h2>结论</h2>
<div class="box">
<p>在持两只实物美债：价格端 <span class="{cls(s['open_price_pnl'])}">${money(s['open_price_pnl'])}</span>，
票息 <span class="pos">${money(s['open_coupons'])}</span>，综合
<span class="{cls(s['open_total_pnl'])}">${money(s['open_total_pnl'])}</span>。</p>
<p>计入已结清链路后，美债兑换全链路综合约
<span class="{cls(s['all_in_pnl'])}">${money(s['all_in_pnl'])}</span> USD。</p>
</div>
<div class="foot">Skill: mti-cncbi-ai-analysis &nbsp;|&nbsp; 每周一 11:00 launchd<br/>
生成：{datetime.now():%Y-%m-%d %H:%M:%S}<br/>
内部持仓管理用途，非投资建议。</div>
</body></html>
"""


def write_markdown(report: dict[str, Any], y_src: str, cfg: dict[str, Any]) -> str:
    s = report["summary"]
    o = report["open"]
    lines = [
        f"# AI分析-{report['asof'].replace('-', '')} MTI 美债盈亏（含票息）",
        "",
        f"- 估值日：{report['asof']}",
        f"- US30Y：{report['y30']*100:.3f}%（{y_src}）",
        f"- 账户：{cfg.get('account')}",
        f"- 主体：{cfg.get('entity')}",
        "",
        "## 摘要",
        f"- 在持市值：USD {money(s['open_mv'])}",
        f"- 在持净成本：USD {money(s['open_net_cost'])}",
        f"- 价格盈亏：USD {money(s['open_price_pnl'])}",
        f"- 已收票息：USD {money(s['open_coupons'])}",
        f"- **在持综合盈亏：USD {money(s['open_total_pnl'])}**",
        f"- 已结清综合：USD {money(s['closed_total_pnl'])}",
        f"- **全链路综合：USD {money(s['all_in_pnl'])}**",
        "",
        "## 在持分券",
        f"- UC08 4.25% 08/15/54 面值 {money(o['uc08']['face'])} @ {o['uc08']['px']:.3f} → MV {money(o['uc08']['mv'])}；票息 {money(o['uc08']['coupons'])}；综合 {money(o['uc08']['total_pnl'])}",
        f"- UE63 4.50% 11/15/54 面值 {money(o['ue63']['face'])} @ {o['ue63']['px']:.3f} → MV {money(o['ue63']['mv'])}；票息 {money(o['ue63']['coupons'])}；综合 {money(o['ue63']['total_pnl'])}",
        "",
        "## 已结清",
        f"- UG12 100万链路：{money(report['closed']['ug12_1mm_total'])}",
        f"- SN36 50万：{money(report['closed']['ug12_05_total'])}",
        "",
        "## 待确认票息",
    ]
    if report.get("pending_coupons"):
        for p in report["pending_coupons"]:
            lines.append(f"- {p['bond']} {p['date']}: {money(p['amount'])} — {p['note']}")
    else:
        lines.append("- 无")
    lines += [
        "",
        f"生成时间：{datetime.now():%Y-%m-%d %H:%M:%S}",
        "口径：净成本=票据本金−交割零头；价格=Clean price（YTM=US30Y）。",
    ]
    return "\n".join(lines) + "\n"


def html_to_pdf(html_path: Path, pdf_path: Path, logger: Logger) -> bool:
    chrome = Path("/Applications/Google Chrome.app/Contents/MacOS/Google Chrome")
    if chrome.exists():
        try:
            subprocess.run(
                [
                    str(chrome),
                    "--headless",
                    "--disable-gpu",
                    "--no-pdf-header-footer",
                    f"--print-to-pdf={pdf_path}",
                    f"file://{html_path}",
                ],
                check=False,
                timeout=120,
                capture_output=True,
            )
            return pdf_path.exists() and pdf_path.stat().st_size > 1000
        except Exception as e:  # noqa: BLE001
            logger.log(f"chrome pdf failed: {e}")
    return False


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="MTI CNCBI UST weekly AI analysis")
    ap.add_argument("--config", help="Path to config.json")
    ap.add_argument("--asof", help="Valuation date YYYY-MM-DD (default today)")
    ap.add_argument("--y30", type=float, help="Override US30Y decimal, e.g. 0.05181")
    args = ap.parse_args(argv)

    cfg = load_config(Path(args.config) if args.config else None)
    workdir = Path(cfg["workdir"]).expanduser()
    log_dir = Path(cfg["log_dir"]).expanduser()
    logger = Logger(log_dir)
    workdir.mkdir(parents=True, exist_ok=True)

    asof = date.fromisoformat(args.asof) if args.asof else date.today()
    logger.log(f"=== MTI weekly analysis start asof={asof} ===")
    if cfg.get("_config_path"):
        logger.log(f"config: {cfg['_config_path']}")

    if args.y30 is not None:
        y30, y_src = float(args.y30), f"CLI override {float(args.y30)*100:.3f}%"
    else:
        y30, y_src = fetch_y30(float(cfg.get("fallback_y30", FALLBACK_Y30)))
    logger.log(f"yield: {y_src}")

    detail = load_detail_workbook(cfg)
    logger.log(f"detail workbook: {detail.get('ok')} {detail.get('path') or detail.get('error')}")

    out_root = ensure_ai_root(cfg, logger)
    out_dir = out_root / f"AI分析-{asof.strftime('%Y%m%d')}"
    out_dir.mkdir(parents=True, exist_ok=True)

    report = compute(asof, y30)
    report["yield_source"] = y_src
    report["detail_workbook"] = {k: detail.get(k) for k in ("ok", "path", "error")}
    report["generated_at"] = datetime.now().isoformat(timespec="seconds")
    report["skill"] = "mti-cncbi-ai-analysis"
    report["account"] = cfg.get("account")
    report["entity"] = cfg.get("entity")

    (workdir / "ust_full_pnl_with_coupons.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    stamp = asof.strftime("%Y%m%d")
    json_path = out_dir / "ust_full_pnl_with_coupons.json"
    md_path = out_dir / f"AI分析-{stamp}.md"
    html_path = out_dir / f"AI分析-{stamp}.html"
    pdf_path = out_dir / f"AI分析-{stamp}.pdf"
    readme = out_dir / "README.txt"

    json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    md_path.write_text(write_markdown(report, y_src, cfg), encoding="utf-8")
    html_path.write_text(render_html(report, y_src, cfg, detail), encoding="utf-8")
    pdf_ok = html_to_pdf(html_path, pdf_path, logger)

    readme.write_text(
        "\n".join(
            [
                f"MTI AI分析 {asof.isoformat()}",
                f"全链路综合盈亏 USD {money(report['summary']['all_in_pnl'])}",
                f"在持综合 USD {money(report['summary']['open_total_pnl'])}",
                f"US30Y {report['y30']*100:.3f}% ({y_src})",
                f"PDF: {'yes' if pdf_ok else 'no'}",
                f"生成: {report['generated_at']}",
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    (out_root / "LATEST.txt").write_text(f"{out_dir.name}\n{report['generated_at']}\n", encoding="utf-8")

    # optional baseline copies from workdir
    for name in (
        "MTI_美债盈亏_含票息_陈艳平.pdf",
        "MTI_美债盈亏_含票息_陈艳平.html",
    ):
        src = workdir / name
        if src.exists():
            shutil.copy2(src, out_dir / f"baseline_{name}")

    logger.log(f"wrote {out_dir}")
    logger.log(
        f"summary open_total={report['summary']['open_total_pnl']} "
        f"all_in={report['summary']['all_in_pnl']} pdf={pdf_ok}"
    )
    logger.log("=== done ===")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception:  # noqa: BLE001
        DEFAULT_LOG_DIR.mkdir(parents=True, exist_ok=True)
        err = traceback.format_exc()
        print(err)
        with (DEFAULT_LOG_DIR / "weekly.err.log").open("a", encoding="utf-8") as f:
            f.write(err + "\n")
        raise
